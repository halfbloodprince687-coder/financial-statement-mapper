"""
Core engine for the Financial Statement Mapper.

Responsibilities
-----------------
1. Read .xls (legacy BIFF) and .xlsx workbooks into plain value grids.
2. Parse a "destination" template workbook: every sheet is expected to carry
   a header row of the form FieldPLBS01, FieldPLBS02, ... marking which
   columns hold machine-readable Field IDs. The cell immediately to the
   LEFT of a Field ID cell is where the numeric value must be written.
3. Parse a "source" workbook (a real financial statement) into a flat
   "ledger" of labelled line items, each carrying every numeric value found
   on that row together with whatever column header (year / period) applies
   to it. This does not assume any fixed layout - it copes with 2-column
   (current/previous) and multi-column (e.g. non-current vs current
   maturities) note tables alike.
4. Fuzzy-match destination fields to source ledger rows so the UI can offer
   a ranked shortlist instead of forcing the user to hunt through 700+ rows.

Nothing here talks to Streamlit - this module is pure data plumbing so it
can be unit-tested / reused headlessly.
"""

from __future__ import annotations

import re
import difflib
from dataclasses import dataclass, field
from typing import Optional


# --------------------------------------------------------------------------
# Grid loading (unifies .xls and .xlsx into the same in-memory structure)
# --------------------------------------------------------------------------

def load_workbook_grids(path: str) -> dict[str, list[list]]:
    """Return {sheet_name: [[cell, cell, ...], ...]} for any .xls/.xlsx file."""
    lower = path.lower()
    if lower.endswith(".xls"):
        return _load_xls(path)
    return _load_xlsx(path)


def _load_xlsx(path: str) -> dict[str, list[list]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    grids = {}
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows():
            rows.append([c.value for c in row])
        grids[ws.title] = rows
    return grids


def _load_xls(path: str) -> dict[str, list[list]]:
    import xlrd

    book = xlrd.open_workbook(path)
    grids = {}
    for sheet in book.sheets():
        rows = []
        for r in range(sheet.nrows):
            row_vals = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                v = cell.value
                if v == "":
                    v = None
                row_vals.append(v)
            rows.append(row_vals)
        grids[sheet.name] = rows
    return grids


def col_letter(idx0: int) -> str:
    """0-based column index -> Excel column letter."""
    n = idx0 + 1
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


# --------------------------------------------------------------------------
# Destination template parsing
# --------------------------------------------------------------------------

FIELD_HEADER_RE = re.compile(r"^FieldPLBS\d+$", re.IGNORECASE)
# A genuine Field ID is a short, space-free identifier like "dom_manufracture"
# or "EQL_sharecap" - this excludes ordinary header/label text like "Particulars".
ID_LOOKS_LIKE_RE = re.compile(r"^[A-Za-z][A-Za-z0-9_]{1,40}$")
# Plain English header/label words that can slip through the identifier regex
# (e.g. "Particulars") but are never actually a Field ID.
EXCLUDE_ID_WORDS = {
    "particulars", "total", "note", "notes", "rs", "amount", "nil", "yes", "no",
    "figures", "current", "previous", "period", "date", "sr", "sno", "percent",
}


@dataclass
class DestField:
    sheet: str
    row0: int                 # 0-based row index
    id_col0: int               # 0-based column index of the Field ID cell
    value_col0: int            # 0-based column index of the value cell (id_col0 - 1)
    field_id: str
    label: str                 # nearest row label, for display / matching
    section: str                # nearest section heading (e.g. "EQUITY AND LIABILITIES")
    existing_value: object      # whatever is already in the value cell (usually None or 0)
    is_total: bool = False
    scale_default: bool = True  # whether the multiplication factor should apply by default


NON_SCALE_HINTS = [
    "earning per", "earnings per", "eps", "per share", "% ", "percent",
    "no. of shares", "number of shares", "no of shares", "% holding",
    "yes (1)", "/ no", "ratio", "multiple", "x)", "basic", "diluted",
]


def _guess_scale_default(label: str) -> bool:
    low = label.lower()
    return not any(h in low for h in NON_SCALE_HINTS)


def parse_destination(grids: dict[str, list[list]]) -> dict[str, list[DestField]]:
    result: dict[str, list[DestField]] = {}
    for sheet, rows in grids.items():
        header_row0 = None
        for r0, row in enumerate(rows[:10]):
            hits = sum(1 for v in row if isinstance(v, str) and FIELD_HEADER_RE.match(v.strip()))
            if hits >= 1:
                header_row0 = r0
                break
        if header_row0 is None:
            continue  # not a Field-ID driven sheet

        id_cols_raw = [c for c, v in enumerate(rows[header_row0])
                       if isinstance(v, str) and FIELD_HEADER_RE.match(v.strip())]

        # The header row tags several columns as "FieldPLBSxx", but in
        # practice only ONE of them (per value/id pair) actually holds
        # machine-readable IDs on data rows - the others are the ordinary
        # label column, which occasionally contains a single bare word
        # ("Royalty", "Deposits", "Diluted", "ASSETS") that would otherwise
        # look like a plausible Field ID. Decide per-column, from the data
        # itself: a genuine ID column is overwhelmingly single-token
        # (no spaces); a label column is mostly multi-word phrases.
        # Column 0 can never be an ID column either way - there is no cell
        # to its left to hold a value.
        id_cols = []
        for c in id_cols_raw:
            if c == 0:
                continue
            samples = [row[c].strip() for row in rows[header_row0 + 1:]
                       if c < len(row) and isinstance(row[c], str) and row[c].strip()]
            if not samples:
                continue
            space_frac = sum(1 for v in samples if " " in v) / len(samples)
            if space_frac > 0.15:
                continue  # looks like the descriptive label column, not IDs
            id_cols.append(c)

        fields: list[DestField] = []
        current_section = sheet
        for r0 in range(header_row0 + 1, len(rows)):
            row = rows[r0]
            # Track section headings: a row where column B (idx1) has long text
            # and no field id present on the row.
            label_cell = None
            for c in range(min(2, len(row))):
                if isinstance(row[c], str) and row[c].strip():
                    label_cell = row[c].strip()
                    break
            row_has_id = False
            for idc in id_cols:
                if (idc < len(row) and isinstance(row[idc], str)
                        and ID_LOOKS_LIKE_RE.match(row[idc].strip())
                        and row[idc].strip().lower() not in EXCLUDE_ID_WORDS):
                    value_col0 = idc - 1
                    if value_col0 < 0:
                        continue  # no cell to write the value into - skip
                    row_has_id = True
                    fid = row[idc].strip()
                    existing = row[value_col0] if value_col0 < len(row) else None
                    lbl = label_cell or fid
                    is_total = "total" in lbl.lower()
                    fields.append(DestField(
                        sheet=sheet, row0=r0, id_col0=idc, value_col0=value_col0,
                        field_id=fid, label=lbl, section=current_section,
                        existing_value=existing, is_total=is_total,
                        scale_default=_guess_scale_default(lbl),
                    ))
            if not row_has_id and label_cell and len(label_cell) > 3 and label_cell.isupper() is False:
                # heuristic: an unnumbered, id-less text row often marks a section
                if label_cell[:1].isalpha() or label_cell[:1] in "IVX":
                    current_section = label_cell[:80]

        result[sheet] = fields
    return result


# --------------------------------------------------------------------------
# Source ledger parsing
# --------------------------------------------------------------------------

YEAR_RE = re.compile(r"(19|20)\d{2}\s*[-/]\s*(\d{2}|(19|20)\d{2})")
PERIOD_KEYWORDS = ["current reporting period", "previous reporting period",
                   "non-current portion", "current maturities",
                   "figures as at", "figures for the period"]


def _header_strength(text: str) -> bool:
    if not isinstance(text, str):
        return False
    t = text.lower()
    if YEAR_RE.search(t):
        return True
    return any(k in t for k in PERIOD_KEYWORDS)


@dataclass
class SourceItem:
    sheet: str
    row0: int
    label: str
    breadcrumb: str            # "Note 5 Long-term borrowings"
    values: dict = field(default_factory=dict)   # {header_label_or_colletter: value}

    def search_text(self) -> str:
        return f"{self.breadcrumb} {self.label}".lower()


NUMERIC_TYPES = (int, float)


def parse_source(grids: dict[str, list[list]]) -> list[SourceItem]:
    items: list[SourceItem] = []
    for sheet, rows in grids.items():
        active_headers: dict[int, str] = {}
        excluded_cols: set[int] = set()
        current_note = ""
        for r0, row in enumerate(rows):
            # column(s) labelled "Note No." (or similar) are references, not amounts
            for c, v in enumerate(row):
                if isinstance(v, str) and "note no" in v.lower():
                    excluded_cols.add(c)

            # 1) does this row refresh the column headers (year / period labels)?
            row_headers = {c: v.strip() for c, v in enumerate(row)
                           if isinstance(v, str) and _header_strength(v)}
            if row_headers:
                active_headers.update(row_headers)
                # a pure header row rarely also carries data worth extracting
                if all((not isinstance(v, NUMERIC_TYPES)) for v in row):
                    continue

            # 2) does this row start a new "Note N  Title" block?
            if (len(row) > 1 and isinstance(row[0], (int, float))
                    and isinstance(row[1], str) and row[1].strip()
                    and float(row[0]).is_integer() and 0 < row[0] < 100):
                current_note = f"Note {int(row[0])} {row[1].strip()}"

            # 3) extract label (first non-empty text cell, scanning left->right,
            #    skipping pure section markers like single letters/roman numerals
            #    only if a longer label exists later in the row)
            label = None
            label_col = None
            for c, v in enumerate(row):
                if isinstance(v, str) and v.strip() and not _header_strength(v):
                    label = v.strip()
                    label_col = c
                    break
            if not label or len(label) < 2:
                continue
            if label.lower() in ("particulars", "amount (in rs.)", "rs.", "amount in rs."):
                continue

            # 4) collect numeric cells after the label column
            values = {}
            for c, v in enumerate(row):
                if c <= (label_col or -1):
                    continue
                if c in excluded_cols:
                    continue
                if isinstance(v, NUMERIC_TYPES):
                    hdr = active_headers.get(c, f"col {col_letter(c)}")
                    values[hdr] = v
            if not values:
                continue

            items.append(SourceItem(
                sheet=sheet, row0=r0, label=label,
                breadcrumb=current_note, values=values,
            ))
    return items


# --------------------------------------------------------------------------
# Fuzzy matching
# --------------------------------------------------------------------------

def _normalize(text: str) -> str:
    t = text.lower()
    t = re.sub(r"[^a-z0-9 ]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


@dataclass
class LogicalRow:
    """One or more DestField slots that should be filled from the SAME source
    line item - typically a Current/Previous-year pair that lives on two
    different destination sheets but represents one real-world figure."""
    key: str
    label: str
    section: str
    slots: list[DestField]   # ordered: first = most-current period, then older


# Sheet pairs where row-for-row structure is identical and a single source
# match should populate both the "Current" and "Previous" period sheets.
SHEET_PAIRS = [("PLCurrent", "PLPrevious"), ("BSCurrent", "BSPrevious")]
STANDALONE_SHEETS = ["BSTrade", "Financial", "Capital"]


def build_logical_rows(dest_fields: dict[str, list[DestField]]) -> list[LogicalRow]:
    rows: list[LogicalRow] = []

    for cur_sheet, prev_sheet in SHEET_PAIRS:
        cur_fields = dest_fields.get(cur_sheet, [])
        prev_fields = dest_fields.get(prev_sheet, [])
        prev_by_row: dict[int, list[DestField]] = {}
        for f in prev_fields:
            prev_by_row.setdefault(f.row0, []).append(f)
        prev_used: dict[int, int] = {}
        for f in cur_fields:
            slots = [f]
            candidates = prev_by_row.get(f.row0, [])
            used = prev_used.get(f.row0, 0)
            if used < len(candidates):
                slots.append(candidates[used])
                prev_used[f.row0] = used + 1
            rows.append(LogicalRow(
                key=f"{cur_sheet}:{f.row0}:{f.id_col0}", label=f.label, section=f.section, slots=slots,
            ))

    for sheet in STANDALONE_SHEETS:
        for f in dest_fields.get(sheet, []):
            rows.append(LogicalRow(
                key=f"{sheet}:{f.row0}:{f.id_col0}", label=f.label, section=f.section, slots=[f],
            ))

    return rows


def sorted_value_keys(values: dict) -> list[str]:
    """Order a source item's {header: value} dict from most-recent period to
    oldest, so slot[0] (Current sheet) gets the newest figure and slot[1]
    (Previous sheet) gets the next one. Falls back to insertion order for
    headers where no year can be parsed."""
    def year_of(k: str) -> Optional[int]:
        m = YEAR_RE.search(k)
        if not m:
            return None
        y = m.group(0)
        digits = re.findall(r"\d+", y)
        first = int(digits[0])
        if first < 100:
            first += 2000
        return first

    keys = list(values.keys())
    with_year = [(k, year_of(k)) for k in keys]
    if any(y is not None for _, y in with_year):
        with_year.sort(key=lambda t: (t[1] is None, -(t[1] or 0)))
        return [k for k, _ in with_year]
    return keys


def suggest_matches(dest_label: str, ledger: list[SourceItem], top_n: int = 5) -> list[tuple[float, SourceItem]]:
    target = _normalize(dest_label)
    target_tags = classify_categories(target)
    scored = []
    for item in ledger:
        cand = _normalize(item.label)
        score = difflib.SequenceMatcher(None, target, cand).ratio()
        # small bonus if every significant word of the dest label appears in candidate
        words = [w for w in target.split() if len(w) > 3]
        if words:
            hits = sum(1 for w in words if w in cand)
            score += 0.15 * (hits / len(words))

        # Hard-disqualify pairs that share most of their words but disagree on
        # a meaning-flipping category - e.g. "Short-term borrowings" and
        # "Long-term borrowings" score high on raw text overlap but are
        # opposite balance-sheet classifications. Plain string similarity
        # cannot tell these apart, so it's enforced explicitly here.
        cand_tags = classify_categories(cand)
        if _categories_conflict(target_tags, cand_tags):
            score *= 0.05

        scored.append((score, item))
    scored.sort(key=lambda x: x[0], reverse=True)
    return scored[:top_n]


# Mutually-exclusive classification groups. For each group, a piece of text is
# tagged with at most one label - order matters, since e.g. "non current"
# must be checked before the bare word "current" (which is a substring of it
# once punctuation is normalized away).
_CATEGORY_GROUPS: list[list[tuple[str, list[str]]]] = [
    [  # maturity / term
        ("LONG", ["non current", "noncurrent", "long term", "longterm"]),
        ("SHORT", ["short term", "shortterm", "current maturities", "current"]),
    ],
    [  # secured vs unsecured
        ("UNSECURED", ["unsecured"]),
        ("SECURED", ["secured"]),
    ],
]


def classify_categories(normalized_text: str) -> dict[str, str]:
    tags: dict[str, str] = {}
    padded = f" {normalized_text} "
    for group_idx, group in enumerate(_CATEGORY_GROUPS):
        for label, phrases in group:
            if any(f" {p} " in padded or padded.startswith(f"{p} ") or padded.endswith(f" {p}")
                   for p in phrases):
                tags[f"group{group_idx}"] = label
                break
    return tags


def _categories_conflict(a: dict[str, str], b: dict[str, str]) -> bool:
    for k, v in a.items():
        if k in b and b[k] != v:
            return True
    return False
