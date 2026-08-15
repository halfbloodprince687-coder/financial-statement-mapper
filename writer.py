"""Writes the reviewed mapping into a copy of the destination .xlsx template.

Cells that already contain a formula (e.g. the template's built-in
``=SUM(...)`` subtotal rows) are left untouched on purpose - overwriting them
with a static number would silently break recalculation the next time the
sheet is opened in Excel.
"""

from __future__ import annotations

import openpyxl


def load_formula_map(xlsx_path: str) -> dict[tuple[str, int, int], str]:
    """Return {(sheet, row0, col0): formula_string} for every cell that
    already holds a formula in the destination template."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    out = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    out[(ws.title, c.row - 1, c.column - 1)] = c.value
    return out


def structure_summary(xlsx_path: str) -> dict:
    """Quick sheet-level summary of a workbook, used to reassure the user the
    export still matches the original template's shape."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    return {
        "sheets": len(wb.worksheets),
        "sheet_names": [ws.title for ws in wb.worksheets],
    }


def write_output(template_xlsx_path: str, out_path: str,
                  cell_values: dict[tuple[str, int, int], float],
                  formula_cells: set[tuple[str, int, int]]) -> list[str]:
    """cell_values: {(sheet, row0, col0): value}. Formula cells are skipped.
    Returns a list of warning strings (e.g. skipped-formula notices).

    Loads the ORIGINAL template workbook and edits it in place before saving
    a copy - every sheet, formula, cell style, merged range, column width
    etc. that isn't one of the mapped value cells comes through unchanged,
    since nothing about the workbook is rebuilt from scratch."""
    warnings: list[str] = []
    wb = openpyxl.load_workbook(template_xlsx_path, data_only=False)
    for (sheet, row0, col0), value in cell_values.items():
        if (sheet, row0, col0) in formula_cells:
            warnings.append(
                f"Skipped {sheet}!R{row0+1}C{col0+1}: already a formula in the template, left untouched."
            )
            continue
        if sheet not in wb.sheetnames:
            warnings.append(f"Sheet '{sheet}' not found in template - skipped.")
            continue
        ws = wb[sheet]
        ws.cell(row=row0 + 1, column=col0 + 1, value=value)
    wb.save(out_path)
    return warnings
